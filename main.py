from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# Load the input workbook and worksheet
input_wb = load_workbook("input.xlsx")
input_ws = input_wb.active

# Create a new workbook and worksheet
output_wb = Workbook()
output_ws = output_wb.active

# Write column headings to output worksheet
output_ws.append(["Question", "Choice 1", "Choice 2", "Choice 3", "Choice 4", "sub 1","sub 2","sub 3", "sub 4" ])

for row in input_ws.iter_rows(min_row=2, values_only=True):
    url = row[0]

    driver = webdriver.Chrome()
    driver.get(url)

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    # Find the question body and extract the text content
    question = soup.find("div", class_="TopprQuestion_questionBody__OBlTc").get_text()

    # Find all four choices and extract the text content
    choices = soup.find_all("div", class_="Option_content__2ZU_b")
    choice_texts = [choice.get_text() for choice in choices]

    # If there are less than 4 choices, fill the remaining choices with "N/A"
    while len(choice_texts) < 4:
        choice_texts.append("N/A")

    #find all breadcrumbs and extract the text content
    subs = soup.find_all("span", class_="Breadcrumb_name__erY_f ellipsis")
    sub_texts = [sub.get_text() for sub in subs]

    # Rearrange choices based on their order on the page
    ordered_choices = [None] * 4
    for i, choice in enumerate(choices):
        choice_text = choice.get_text()
        index = choice_texts.index(choice_text)
        ordered_choices[index] = choice_text

    # Rearrange subs based on their order on the page
    ordered_subs = [None] * 5
    for i, sub in enumerate(subs):
        sub_text = sub.get_text()
        index = sub_texts.index(sub_text)
        ordered_subs[index] = sub_text

    # Write the question, choices, and subs to the output worksheet
    output_ws.append([question] + choice_texts + sub_texts)

    driver.quit()

# Save the output workbook to a file
output_wb.save("output.xlsx")
